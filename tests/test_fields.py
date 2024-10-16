import datetime as dt
from decimal import Decimal
from types import SimpleNamespace

import pytest
from openpyxl.cell import Cell
from openpyxl.worksheet.worksheet import Worksheet
from rest_framework.fields import (
    BooleanField,
    CharField,
    DateField,
    DateTimeField,
    DecimalField,
    FloatField,
    IntegerField,
    ListField,
    TimeField,
)

from drf_excel.fields import (
    XLSXBooleanField,
    XLSXDateField,
    XLSXField,
    XLSXListField,
    XLSXNumberField,
)
from drf_excel.utilities import XLSXStyle


@pytest.fixture
def style():
    return XLSXStyle()


class TestXLSXField:
    def test_init(self, style: XLSXStyle):
        f = XLSXField(
            key="foo",
            value="bar",
            field=CharField(),
            style=style,
            mapping="",
            cell_style=style,
        )
        assert f.key == "foo"
        assert f.original_value == "bar"
        assert f.value == "bar"

    def test_cell(self, style: XLSXStyle, worksheet: Worksheet):
        f = XLSXField(
            key="foo",
            value="bar",
            field=CharField(),
            style=style,
            mapping="",
            cell_style=style,
        )
        cell = f.cell(worksheet, 1, 1)
        assert isinstance(cell, Cell)
        assert cell.value == "bar"

    def test_cell_with_mapping_str(self, style: XLSXStyle, worksheet: Worksheet):
        f = XLSXField(
            key="foo",
            value={"custom": "bar"},
            field=CharField(),
            style=style,
            mapping="custom",
            cell_style=style,
        )
        cell = f.cell(worksheet, 1, 1)
        assert isinstance(cell, Cell)
        assert cell.value == "bar"

    def test_cell_with_mapping_callable(self, style: XLSXStyle, worksheet: Worksheet):
        f = XLSXField(
            key="foo",
            value=SimpleNamespace(custom="bar"),
            field=CharField(),
            style=style,
            mapping=lambda v: v.custom,
            cell_style=style,
        )
        cell = f.cell(worksheet, 1, 1)
        assert isinstance(cell, Cell)
        assert cell.value == "bar"

    def test_cell_with_invalid_mapping(self, style: XLSXStyle, worksheet: Worksheet):
        f = XLSXField(
            key="foo",
            value="bar",
            field=CharField(),
            style=style,
            mapping=55,
            cell_style=style,
        )
        cell = f.cell(worksheet, 1, 1)
        assert isinstance(cell, Cell)
        assert cell.value == "bar"


class TestXLSXNumberField:
    @pytest.mark.parametrize(
        ("original_value", "cleaned_value"),
        [
            # Conversion worked
            (42, 42),
            ("42", 42),
            (42.1, 42),
            (True, 1),
            (False, 0),
            # Conversion fails
            ("42.0", "42.0"),
            ("foo", "foo"),
            (None, None),
        ],
    )
    def test_init_integer_field(self, style: XLSXStyle, original_value, cleaned_value):
        f = XLSXNumberField(
            key="age",
            value=original_value,
            field=IntegerField(),
            style=style,
            mapping="",
            cell_style=style,
        )
        assert f.key == "age"
        assert f.original_value == original_value
        assert f.value == cleaned_value

    @pytest.mark.parametrize(
        ("original_value", "cleaned_value"),
        [
            # Conversion worked
            (49, 49),
            ("36", 36),
            (11.4, 11.4),
            ("42.0", 42.0),
            (True, 1),
            (False, 0),
            # Conversion fails
            ("foo", "foo"),
            (None, None),
        ],
    )
    def test_init_float_field(self, style: XLSXStyle, original_value, cleaned_value):
        f = XLSXNumberField(
            key="temperature",
            value=original_value,
            field=FloatField(),
            style=style,
            mapping="",
            cell_style=style,
        )
        assert f.key == "temperature"
        assert f.original_value == original_value
        assert f.value == cleaned_value

    @pytest.mark.parametrize(
        ("original_value", "cleaned_value"),
        [
            # Conversion worked
            (49, Decimal(49)),
            ("36", Decimal(36)),
            (11.4, Decimal(11.4)),
            ("42.20", Decimal("42.20")),
            (True, Decimal(1)),
            (False, Decimal(0)),
            (Decimal(5), Decimal(5)),
            (Decimal("35.00"), Decimal("35.00")),
            (Decimal("55.9"), Decimal("55.9")),
            (Decimal("123.9999"), Decimal("123.9999")),
            # Conversion fails
            ("foo", "foo"),
            (None, None),
        ],
    )
    def test_init_decimal_field(self, style: XLSXStyle, original_value, cleaned_value):
        f = XLSXNumberField(
            key="price",
            value=original_value,
            field=DecimalField(max_digits=10, decimal_places=2),
            style=style,
            mapping="",
            cell_style=style,
        )
        assert f.key == "price"
        assert f.original_value == original_value
        assert f.value == cleaned_value

    def test_cell_integer(self, style: XLSXStyle, worksheet: Worksheet):
        f = XLSXNumberField(
            key="age",
            value=42,
            field=IntegerField(),
            style=style,
            mapping="",
            cell_style=style,
        )
        cell = f.cell(worksheet, 1, 1)
        assert isinstance(cell, Cell)
        assert cell.value == 42
        assert cell.number_format == "0"

    def test_cell_float(self, style: XLSXStyle, worksheet: Worksheet):
        f = XLSXNumberField(
            key="weight",
            value=35.5,
            field=FloatField(),
            style=style,
            mapping="",
            cell_style=style,
        )
        cell = f.cell(worksheet, 1, 1)
        assert isinstance(cell, Cell)
        assert cell.value == 35.5
        assert cell.number_format == "0.00"


class TestXLSXDateField:
    @pytest.mark.parametrize(
        ("original_value", "cleaned_value"),
        [
            # Successful parsing
            ("2020-01-01", dt.datetime(2020, 1, 1, 0, 0)),
            ("2020-02-03T13:01", dt.datetime(2020, 2, 3, 13, 1)),
            ("2020-03-04T13:15:30", dt.datetime(2020, 3, 4, 13, 15, 30)),
            ("2020-10-08T15:18:23Z", dt.datetime(2020, 10, 8, 15, 18, 23)),
            # Failed parsing
            ("foo", "foo"),
            (None, None),
            (True, True),
            (300, 300),
            (99.99, 99.99),
        ],
    )
    def test_init_datetime_field(self, style: XLSXStyle, original_value, cleaned_value):
        f = XLSXDateField(
            key="dt",
            value=original_value,
            field=DateTimeField(),
            style=style,
            mapping="",
            cell_style=style,
        )
        assert f.original_value == original_value
        assert f.value == cleaned_value

    @pytest.mark.parametrize(
        ("original_value", "cleaned_value"),
        [
            # Successful parsing
            ("2020-01-01", dt.date(2020, 1, 1)),
            # Failed parsing
            ("2020-02-03T13:01", None),
            ("2020-03-04T13:15:30", None),
            ("2020-10-08T15:18:23.1234Z", None),
            ("foo", None),
            (None, None),
            (True, True),
            (300, 300),
            (99.99, 99.99),
        ],
    )
    def test_init_date_field(self, style: XLSXStyle, original_value, cleaned_value):
        f = XLSXDateField(
            key="d",
            value=original_value,
            field=DateField(),
            style=style,
            mapping="",
            cell_style=style,
        )
        assert f.original_value == original_value
        assert f.value == cleaned_value

    @pytest.mark.parametrize(
        ("original_value", "cleaned_value"),
        [
            # Successful parsing
            ("13:01", dt.time(13, 1)),
            ("13:15:30", dt.time(13, 15, 30)),
            # Failed parsing
            ("2020-01-01", "2020-01-01"),
            ("2020-02-03T13:01", "2020-02-03T13:01"),
            ("2020-03-04T13:15:30", "2020-03-04T13:15:30"),
            ("2020-10-08T15:18:23.1234Z", "2020-10-08T15:18:23.1234Z"),
            ("foo", "foo"),
            (None, None),
            (True, True),
            (300, 300),
            (99.99, 99.99),
        ],
    )
    def test_init_time_field(self, style: XLSXStyle, original_value, cleaned_value):
        f = XLSXDateField(
            key="t",
            value=original_value,
            field=TimeField(),
            style=style,
            mapping="",
            cell_style=style,
        )
        assert f.original_value == original_value
        assert f.value == cleaned_value

    def test_cell_datetime_default_format(self, style: XLSXStyle, worksheet: Worksheet):
        f = XLSXDateField(
            key="dt",
            value="2019-03-04T13:15:30",
            field=DateTimeField(),
            style=style,
            mapping="",
            cell_style=style,
        )
        cell = f.cell(worksheet, 1, 1)
        assert isinstance(cell, Cell)
        assert cell.value == dt.datetime(2019, 3, 4, 13, 15, 30)
        assert cell.number_format == "yyyy-mm-dd h:mm:ss"

    def test_cell_datetime_custom_format(
        self, style: XLSXStyle, worksheet: Worksheet, settings
    ):
        settings.DRF_EXCEL_DATETIME_FORMAT = "dd/mm/yyyy h:mm:ss"
        f = XLSXDateField(
            key="dt",
            value="01-04-2015 05:16:09",
            field=DateTimeField(format="%d-%m-%Y %H:%M:%S"),
            style=style,
            mapping="",
            cell_style=style,
        )
        cell = f.cell(worksheet, 1, 1)
        assert isinstance(cell, Cell)
        assert cell.value == dt.datetime(2015, 4, 1, 5, 16, 9)
        assert cell.number_format == "dd/mm/yyyy h:mm:ss"

    def test_cell_date_default_format(self, style: XLSXStyle, worksheet: Worksheet):
        f = XLSXDateField(
            key="dt",
            value="2018-09-10",
            field=DateField(),
            style=style,
            mapping="",
            cell_style=style,
        )
        cell = f.cell(worksheet, 1, 1)
        assert isinstance(cell, Cell)
        assert cell.value == dt.date(2018, 9, 10)
        assert cell.number_format == "yyyy-mm-dd"

    def test_cell_date_custom_format(
        self, style: XLSXStyle, worksheet: Worksheet, settings
    ):
        settings.DRF_EXCEL_DATE_FORMAT = "dd/mm/yyyy"
        f = XLSXDateField(
            key="dt",
            value="25/10/2017",
            field=DateField(format="%d/%m/%Y"),
            style=style,
            mapping="",
            cell_style=style,
        )
        cell = f.cell(worksheet, 1, 1)
        assert isinstance(cell, Cell)
        assert cell.value == dt.date(2017, 10, 25)
        assert cell.number_format == "dd/mm/yyyy"

    def test_cell_time_default_format(self, style: XLSXStyle, worksheet: Worksheet):
        f = XLSXDateField(
            key="dt",
            value="09:10:11",
            field=TimeField(),
            style=style,
            mapping="",
            cell_style=style,
        )
        cell = f.cell(worksheet, 1, 1)
        assert isinstance(cell, Cell)
        assert cell.value == dt.time(9, 10, 11)
        assert cell.number_format == "h:mm:ss"

    def test_cell_time_custom_format(
        self, style: XLSXStyle, worksheet: Worksheet, settings
    ):
        settings.DRF_EXCEL_TIME_FORMAT = "hh - mm - ss"
        f = XLSXDateField(
            key="dt",
            value="09h 10m 11s",
            field=TimeField(format="%Hh %Mm %Ss"),
            style=style,
            mapping="",
            cell_style=style,
        )
        cell = f.cell(worksheet, 1, 1)
        assert isinstance(cell, Cell)
        assert cell.value == dt.time(9, 10, 11)
        assert cell.number_format == "hh - mm - ss"


class TestXLSXListField:
    def test_cell_default_separator(self, style: XLSXStyle, worksheet: Worksheet):
        f = XLSXListField(
            list_sep=None,
            key="things",
            value=["foo", "bar", "baz"],
            field=ListField(),
            style=style,
            mapping="",
            cell_style=style,
        )
        assert f.original_value == f.value == ["foo", "bar", "baz"]
        cell = f.cell(worksheet, 1, 1)
        assert isinstance(cell, Cell)
        assert cell.value == "foo, bar, baz"

    def test_cell_custom_separator(self, style: XLSXStyle, worksheet: Worksheet):
        f = XLSXListField(
            list_sep=";",
            key="things",
            value=["john", "doe", "john.doe@example.com"],
            field=ListField(),
            style=style,
            mapping="",
            cell_style=style,
        )
        assert f.original_value == f.value == ["john", "doe", "john.doe@example.com"]
        cell = f.cell(worksheet, 1, 1)
        assert isinstance(cell, Cell)
        assert cell.value == "john;doe;john.doe@example.com"

    def test_cell_complex_types(self, style: XLSXStyle, worksheet: Worksheet):
        f = XLSXListField(
            list_sep=None,
            key="objs",
            value=[{"a": 1}, {"b": 2}],
            field=ListField(),
            style=style,
            mapping="",
            cell_style=style,
        )
        assert f.original_value == f.value == [{"a": 1}, {"b": 2}]
        cell = f.cell(worksheet, 1, 1)
        assert isinstance(cell, Cell)
        assert cell.value == '[{"a": 1}, {"b": 2}]'

    def test_cell_with_empty_list(self, style: XLSXStyle, worksheet: Worksheet):
        f = XLSXListField(
            list_sep=None,
            key="objs",
            value=[],
            field=ListField(),
            style=style,
            mapping="",
            cell_style=style,
        )
        assert f.original_value == f.value == []
        cell = f.cell(worksheet, 1, 1)
        assert isinstance(cell, Cell)
        assert cell.value == ""

    def test_cell_with_null_value(self, style: XLSXStyle, worksheet: Worksheet):
        f = XLSXListField(
            list_sep=None,
            key="objs",
            value=None,
            field=ListField(allow_null=True),
            style=style,
            mapping="",
            cell_style=style,
        )
        assert f.original_value is f.value is None
        cell = f.cell(worksheet, 1, 1)
        assert isinstance(cell, Cell)
        assert cell.value is None


class TestXLSXBooleanField:
    @pytest.mark.parametrize("value", [True, False])
    def test_cell_default_display(
        self, style: XLSXStyle, worksheet: Worksheet, value: bool
    ):
        f = XLSXBooleanField(
            boolean_display={},
            key="is_active",
            value=value,
            field=BooleanField(),
            style=style,
            mapping="",
            cell_style=style,
        )
        cell = f.cell(worksheet, 1, 1)
        assert isinstance(cell, Cell)
        assert cell.value == value

    @pytest.mark.parametrize(
        ("original_value", "cleaned_value"),
        [
            (True, "yes"),
            (False, "no"),
        ],
    )
    def test_cell_custom_display(
        self,
        style: XLSXStyle,
        worksheet: Worksheet,
        original_value: bool,
        cleaned_value: str,
    ):
        f = XLSXBooleanField(
            boolean_display={True: "yes", False: "no"},
            key="is_active",
            value=original_value,
            field=BooleanField(),
            style=style,
            mapping="",
            cell_style=style,
        )
        cell = f.cell(worksheet, 1, 1)
        assert isinstance(cell, Cell)
        assert cell.value == cleaned_value
