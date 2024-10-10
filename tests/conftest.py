import io
from typing import Union, Callable

import pytest
from openpyxl.reader.excel import load_workbook
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet


@pytest.fixture
def workbook() -> Workbook:
    return Workbook()


@pytest.fixture
def worksheet(workbook: Workbook) -> Worksheet:
    return Worksheet(workbook)


@pytest.fixture
def workbook_reader() -> Callable[[Union[bytes, str]], Workbook]:
    def reader_func(buffer: Union[bytes, str]) -> Workbook:
        io_buffer = io.BytesIO(buffer)
        return load_workbook(io_buffer, read_only=True)

    return reader_func
