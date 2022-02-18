import json
from collections.abc import Iterable, MutableMapping
from typing import Dict

from openpyxl import Workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import Alignment, Border, Font, NamedStyle, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.writer.excel import save_virtual_workbook
from rest_framework.fields import (
    BooleanField,
    DateField,
    DateTimeField,
    DecimalField,
    Field,
    FloatField,
    IntegerField,
    ListField,
    NullBooleanField,
    TimeField,
)
from rest_framework.renderers import BaseRenderer
from rest_framework.serializers import Serializer

from drf_renderer_xlsx.fields import XLSXBooleanField, XLSXDateField, XLSXField, XLSXListField, XLSXNumberField


def get_style_from_dict(style_dict, style_name):
    """
    Make NamedStyle instance from dictionary
    :param style_dict: dictionary with style properties.
           Example:    {'fill': {'fill_type'='solid',
                                 'start_color'='FFCCFFCC'},
                        'alignment': {'horizontal': 'center',
                                      'vertical': 'center',
                                      'wrapText': True,
                                      'shrink_to_fit': True},
                        'border_side': {'border_style': 'thin',
                                        'color': 'FF000000'},
                        'font': {'name': 'Arial',
                                 'size': 14,
                                 'bold': True,
                                 'color': 'FF000000'}
                        }
    :param style_name: name of created style
    :return: openpyxl.styles.NamedStyle instance
    """
    style = NamedStyle(name=style_name)
    if not style_dict:
        return style
    for key, value in style_dict.items():
        if key == "font":
            style.font = Font(**value)
        elif key == "fill":
            style.fill = PatternFill(**value)
        elif key == "alignment":
            style.alignment = Alignment(**value)
        elif key == "border_side":
            side = Side(**value)
            style.border = Border(left=side, right=side, top=side, bottom=side)

    return style


def get_attribute(get_from, prop_name, default=None):
    """
    Get attribute from object with name <prop_name>, or take it from function get_<prop_name>
    :param get_from: instance of object
    :param prop_name: name of attribute (str)
    :param default: what to return if attribute doesn't exist
    :return: value of attribute <prop_name> or default
    """
    prop = getattr(get_from, prop_name, None)
    if not prop:
        prop_func = getattr(get_from, "get_{}".format(prop_name), None)
        if prop_func:
            prop = prop_func()
    if prop is None:
        prop = default
    return prop


class XLSXRenderer(BaseRenderer):
    """
    Renderer for Excel spreadsheet open data format (xlsx).
    """

    media_type = "application/xlsx"
    format = "xlsx"
    combined_header_dict = {}
    fields_dict = {}
    ignore_headers = []
    boolean_display = None
    date_format_mappings = None
    number_format_mappings = None
    custom_mappings = None
    custom_cols = None
    sanitize_fields = True  # prepend possibly malicious values with "'"
    list_sep = ", "
    body_style = None

    def render(self, data, accepted_media_type=None, renderer_context=None):
        """
        Render `data` into XLSX workbook, returning a workbook.
        """
        if not self._check_validation_data(data):
            return json.dumps(data)

        if data is None:
            return bytes()

        wb = Workbook()
        self.ws = wb.active

        results = data["results"] if "results" in data else data

        # Take header and column_header params from view
        header = get_attribute(renderer_context["view"], "header", {})
        use_header = header and header.get("use_header", True)
        self.ws.title = header.get("tab_title", "Report")
        header_title = header.get("header_title", "Report")
        img_addr = header.get("img")
        if img_addr:
            img = Image(img_addr)
            self.ws.add_image(img, "A1")
        header_style = get_style_from_dict(header.get("style"), "header_style")

        column_header = get_attribute(renderer_context["view"], "column_header", {})
        column_header_style = get_style_from_dict(column_header.get("style"), "column_header_style")
        column_count = 0
        row_count = 1
        if use_header:
            row_count += 1
        # Make column headers
        column_titles = column_header.get("titles", [])

        # If we have results, get view and serializer from context, then flatten field
        # names
        if len(results):
            drf_view = renderer_context.get("view")

            # Set `xlsx_use_labels = True` inside the API View to enable labels.
            use_labels = getattr(drf_view, "xlsx_use_labels", False)

            # A list of header keys to ignore in our export
            self.ignore_headers = getattr(drf_view, "xlsx_ignore_headers", [])

            # Create a mapping dict named `xlsx_boolean_labels` inside the API View.
            # I.e.: xlsx_boolean_labels: {True: "Yes", False: "No"}
            self.boolean_display = getattr(drf_view, "xlsx_boolean_labels", None)

            # Set dict named xlsx_date_format_mappings with headers as keys and
            # formatting as value. i.e. { 'created_at': 'yyyy-mm-dd h:mm:ss' }
            self.date_format_mappings = getattr(drf_view, "xlsx_date_format_mappings", dict())

            # Set dict named xlsx_number_format_mappings with headers as keys and
            # formatting as value. i.e. { 'cost': '"$"#,##0.00_-' }
            self.number_format_mappings = getattr(drf_view, "xlsx_number_format_mappings", dict())

            # Set dict of additional columns. Can be useful when wanting to add columns
            # that don't exist in the API response. For example, you could want to
            # show values of a dict in individual cols. Takes key, an optional label
            # and value than can be callable
            # Example:
            # {"Additional Col": { label: "Something (optional)", formatter: my_function }}
            self.custom_cols = getattr(drf_view, "xlsx_custom_cols", dict())

            # Map a specific key to a column (I.e. if the field returns a json) or pass
            # a function to format the value
            # Example with key:
            # {"custom_choice": "display"}, showing 'display' in the
            # 'custom_choice' col
            # Example with function:
            # {"custom_choice": custom_func }, passing the value of 'custom_choice' to
            # 'custom_func', allowing for formatting logic
            self.custom_mappings = getattr(drf_view, "xlsx_custom_mappings", dict())

            self.fields_dict = self._serializer_fields(drf_view.get_serializer())

            xlsx_header_dict = self._flatten_serializer_keys(drf_view.get_serializer(), use_labels=use_labels)
            if self.custom_cols:
                custom_header_dict = {
                    key: self.custom_cols[key].get("label", None) or key for key in self.custom_cols.keys()
                }
                self.combined_header_dict = dict(list(xlsx_header_dict.items()) + list(custom_header_dict.items()))
            else:
                self.combined_header_dict = xlsx_header_dict

            for column_name, column_label in self.combined_header_dict.items():
                if column_name == "row_color":
                    continue
                column_count += 1
                if column_count > len(column_titles):
                    column_name_display = column_label
                else:
                    column_name_display = column_titles[column_count - 1]

                self.ws.cell(row=row_count, column=column_count, value=column_name_display).style = column_header_style
            self.ws.row_dimensions[row_count].height = column_header.get("height", 45)

        # Set the header row
        if use_header:
            last_col_letter = "G"
            if column_count:
                last_col_letter = get_column_letter(column_count)
            self.ws.merge_cells("A1:{}1".format(last_col_letter))

            cell = self.ws.cell(row=1, column=1, value=header_title)
            cell.style = header_style
            self.ws.row_dimensions[1].height = header.get("height", 45)

        # Set column width
        column_width = column_header.get("column_width", 20)
        if isinstance(column_width, list):
            for i, width in enumerate(column_width):
                col_letter = get_column_letter(i + 1)
                self.ws.column_dimensions[col_letter].width = width
        else:
            for ws_column in range(1, column_count + 1):
                col_letter = get_column_letter(ws_column)
                self.ws.column_dimensions[col_letter].width = column_width

        # Make body
        body = get_attribute(renderer_context["view"], "body", {})
        self.body_style = get_style_from_dict(body.get("style"), "body_style")
        if isinstance(results, dict):
            self._make_body(body, results, row_count)
        elif isinstance(results, list):
            for row in results:
                self._make_body(body, row, row_count)
                row_count += 1

        return save_virtual_workbook(wb)

    def _check_validation_data(self, data):
        detail_key = "detail"
        if detail_key in data:
            return False
        return True

    def _serializer_fields(self, serializer, parent_key="", key_sep="."):
        _fields_dict = {}
        for k, v in serializer.get_fields().items():
            new_key = f"{parent_key}{key_sep}{k}" if parent_key else k
            if isinstance(v, Serializer):
                _fields_dict.update(self._serializer_fields(v, new_key, key_sep))
            elif isinstance(v, Field):
                _fields_dict[new_key] = v
        return _fields_dict

    def _flatten_serializer_keys(
        self, serializer, parent_key="", parent_label="", key_sep=".", list_sep=", ", label_sep=" > ", use_labels=False
    ):
        """
        Iterate through serializer fields recursively when field is a nested serializer.
        """

        def _get_label(parent_label, label_sep, obj):
            if getattr(v, "label", None):
                if parent_label:
                    return f"{parent_label}{label_sep}{v.label}"
                else:
                    return str(v.label)
            else:
                return False

        _header_dict = {}
        _fields = serializer.get_fields()
        for k, v in _fields.items():
            new_key = f"{parent_key}{key_sep}{k}" if parent_key else k
            # Skip headers we want to ignore
            if new_key in self.ignore_headers:
                continue
            # Iterate through fields if field is a serializer. Check for labels and
            # append if `use_labels` is True. Fallback to using keys.
            if isinstance(v, Serializer):
                if use_labels and getattr(v, "label", None):
                    _header_dict.update(
                        self._flatten_serializer_keys(
                            v, new_key, _get_label(parent_label, label_sep, v), key_sep, list_sep, label_sep, use_labels
                        )
                    )
                else:
                    _header_dict.update(
                        self._flatten_serializer_keys(
                            v, new_key, key_sep=key_sep, list_sep=list_sep, label_sep=label_sep, use_labels=use_labels
                        )
                    )
            elif isinstance(v, Field):
                if use_labels and getattr(v, "label", None):
                    _header_dict[new_key] = _get_label(parent_label, label_sep, v)
                else:
                    _header_dict[new_key] = new_key

        return _header_dict

    def _flatten_data(self, data, parent_key="", key_sep=".") -> Dict[str, XLSXField]:
        items = []
        for k, v in data.items():
            new_key = f"{parent_key}{key_sep}{k}" if parent_key else k
            if isinstance(v, MutableMapping):
                items.extend(self._flatten_data(v, new_key, key_sep=key_sep).items())
            else:
                xlsx_field = self._drf_to_xlsx_field(key=new_key, value=v)
                items.append((new_key, xlsx_field))
        return dict(items)

    def _make_body(self, body, row, row_count):
        column_count = 0
        row_count += 1
        flattened_row = self._flatten_data(row)
        for header_key in self.combined_header_dict:
            if header_key == "row_color":
                continue
            column_count += 1
            field = flattened_row.get(header_key)
            field.cell(self.ws, row_count, column_count) if field else self.ws.cell(row_count, column_count)
        self.ws.row_dimensions[row_count].height = body.get("height", 40)
        if "row_color" in row:
            last_letter = get_column_letter(column_count)
            cell_range = self.ws["A{}".format(row_count) : "{}{}".format(last_letter, row_count)]
            fill = PatternFill(fill_type="solid", start_color=row["row_color"])
            for r in cell_range:
                for c in r:
                    c.fill = fill

    def _drf_to_xlsx_field(self, key, value) -> XLSXField:
        field = self.fields_dict.get(key)
        kwargs = {
            "key": key,
            "value": value,
            "field": field,
            "style": self.body_style,
            # Basically using formatter of custom col as a custom mapping
            "mapping": self.custom_cols.get(key, {}).get("formatter") or self.custom_mappings.get(key),
        }
        date_format = self.date_format_mappings.get(key)
        number_format = self.number_format_mappings.get(key)
        if isinstance(field, BooleanField) or isinstance(field, NullBooleanField) or type(value) == bool:
            return XLSXBooleanField(boolean_display=self.boolean_display, **kwargs)
        elif isinstance(field, IntegerField) or isinstance(field, FloatField) or isinstance(field, DecimalField):
            return XLSXNumberField(number_format=number_format, **kwargs)
        elif isinstance(field, DateTimeField) or isinstance(field, DateField) or isinstance(field, TimeField):
            return XLSXDateField(date_format=date_format, **kwargs)
        elif isinstance(field, ListField) or isinstance(value, Iterable) and not isinstance(value, str):
            return XLSXListField(list_sep=self.list_sep, **kwargs)
        return XLSXField(**kwargs)
